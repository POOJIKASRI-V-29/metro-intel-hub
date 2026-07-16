"""
Excel Extraction module for the KMRL Document Intelligence Platform.

This module leverages openpyxl to process and transform structured tabular data 
from Excel spreadsheets (.xlsx) into semantically enriched text strings suitable 
for embedding generation and vector store ingestion.
"""

import io
import logging
from typing import Any, BinaryIO, List
import openpyxl  # type: ignore
from pydantic import BaseModel, Field

# Setup logger mapping to Stage 0 configurations
logger = logging.getLogger("document_intelligence.ingestion.excel_parser")


class ParsedRow(BaseModel):
    """
    Data model representing a single serialized row from an Excel sheet.
    """
    row_index: int = Field(..., description="The 1-based row index in the worksheet.")
    raw_values: List[str] = Field(..., description="The string representation of each cell value in the row.")
    semantic_text: str = Field(..., description="The flattened, pipe-delimited text layout of the row.")


class ParsedSheet(BaseModel):
    """
    Data model representing an entire parsed worksheet from a workbook.
    """
    sheet_name: str = Field(..., description="The name of the worksheet tab.")
    rows: List[ParsedRow] = Field(..., description="List of parsed data rows extracted from this sheet.")
    combined_markdown: str = Field(..., description="The entire sheet compiled into a Markdown table structure.")


class ExcelParser:
    """
    Handles robust text extraction and semantic flattening from Excel spreadsheets.
    """

    def __init__(self) -> None:
        """Initializes the Excel spreadsheet parser component."""
        pass

    def parse_stream(self, file_stream: BinaryIO) -> List[ParsedSheet]:
        """
        Parses an in-memory binary stream containing an Excel workbook.

        Args:
            file_stream: A seekable file-like binary object containing a spreadsheet.

        Returns:
            A list of ParsedSheet objects containing row elements and markdown conversions.

        Raises:
            ValueError: If the workbook is corrupted or uses an unreadable encryption scheme.
        """
        try:
            file_stream.seek(0)
            stream_bytes = file_stream.read()
            memory_buffer = io.BytesIO(stream_bytes)
            
            # Load workbook using data_only=True to extract evaluated cell values instead of raw formulas
            workbook = openpyxl.load_workbook(memory_buffer, data_only=True, read_only=True)
            return self._extract_content(workbook)

        except Exception as error:
            logger.exception("Failed to parse Excel workbook content from the provided stream.")
            raise ValueError("The provided stream is not a valid or readable Excel workbook.") from error

    def parse_file(self, file_path: str) -> List[ParsedSheet]:
        """
        Parses an Excel workbook file located on the local file system.

        Args:
            file_path: The absolute or relative string path to the target spreadsheet file.

        Returns:
            A list of ParsedSheet objects containing row elements and markdown conversions.

        Raises:
            ValueError: If the file path is invalid, unreadable, or corrupted.
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            return self._extract_content(workbook)

        except FileNotFoundError:
            logger.error(f"Excel file not found at path: {file_path}")
            raise ValueError(f"Spreadsheet file not found at target path: {file_path}")
        except Exception as error:
            logger.exception(f"Structural layout processing failure for Excel file at: {file_path}")
            raise ValueError(f"Corrupt or invalid Excel workbook layout at: {file_path}") from error

    def _extract_content(self, workbook: openpyxl.Workbook) -> List[ParsedSheet]:
        """
        Internal worker that serializes sheets, extracts row tuples, and builds markdown wrappers.

        Args:
            workbook: An openpyxl Workbook instance.

        Returns:
            A list of successfully processed sheet payloads.
        """
        parsed_sheets: List[ParsedSheet] = []

        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_rows: List[ParsedRow] = []
                markdown_lines: List[str] = []
                
                logger.debug(f"Processing Excel worksheet: {sheet_name}")

                for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    # Clean values and convert None cells into clean empty strings
                    row_strings = [str(cell).strip() if cell is not None else "" for cell in row]
                    
                    # Skip rows that are completely empty to optimize downstream processing context
                    if not any(row_strings):
                        continue

                    # Generate semantic metadata layout for the row
                    pipe_delimited = " | ".join(row_strings)
                    semantic_line = f"Sheet: {sheet_name}, Row {r_idx}: {pipe_delimited}"
                    
                    sheet_rows.append(
                        ParsedRow(
                            row_index=r_idx,
                            raw_values=row_strings,
                            semantic_text=semantic_line
                        )
                    )
                    
                    markdown_lines.append(f"| { ' | '.join(row_strings) } |")
                    
                    # Inject markdown table delimiter line directly beneath the first row (assumed header)
                    if len(markdown_lines) == 1:
                        header_separator = "| " + " | ".join(["---"] * len(row_strings)) + " |"
                        markdown_lines.append(header_separator)

                if sheet_rows:
                    combined_md = "\n".join(markdown_lines)
                    parsed_sheets.append(
                        ParsedSheet(
                            sheet_name=sheet_name,
                            rows=sheet_rows,
                            combined_markdown=combined_md
                        )
                    )

            return parsed_sheets

        finally:
            # Clean up unmanaged resources explicitly
            workbook.close()